from django.shortcuts import render, redirect
from ayuda.models import ayuda
from detalle_ayuda.models import Detalle_ayuda
from openpyxl import Workbook
from django.http.response import HttpResponse
from django.views.generic import TemplateView


def home(request):
    context = {'ayuda_list': ayuda.objects.all(), 'detalle_ayuda_list': Detalle_ayuda.objects.all()}
    return render(request, "list.html", context)

class ReporteExcel(TemplateView):
    def gets(self, request, *args, **kwargs):
        Ayuda = ayuda.objects.all()
        Detalle = Detalle_ayuda.objects.all()
        wb = Workbook()
        ws = wb.active
        ws['B1'] = 'Reporte de ayuda'

        ws.merge_cells('B1:H1')

        ws['B3'] = 'Nombre'
        ws['C3'] = 'Beneficiario'
        ws['D3'] = 'Fecha'
        ws['E3'] = 'Entidad'
        ws['F3'] = 'Tipo de ayuda'
        ws['G3'] = 'Detalle'
        ws['H3'] = 'catidad'

        cont = 7

        for ayuda in Ayuda:
            for Detalle_ayuda in Detalle:
                ws.cell(row = cont, column = 2).value = ayuda.nombreAyuda
                ws.cell(row = cont, column = 3).value = ayuda.beneficiario
                ws.cell(row = cont, column = 4).value = ayuda.fecha
                ws.cell(row = cont, column = 5).value = ayuda.entidad
                ws.cell(row = cont, column = 6).value = Detalle_ayuda.tipo_ayuda
                ws.cell(row = cont, column = 7).value = Detalle_ayuda.descripcion
                ws.cell(row = cont, column = 8).value = Detalle_ayuda.cantidad
                cont+=1
        nombre_archivo = "ReporteCovid.xlsx"
        response = HttpResponse(content_type = "aplications/ms-excel")
        content = "attachment; filename = {0}".format(nombre_archivo)
        response['Content-Disposition'] = content
        wb.save(response)
        return response
